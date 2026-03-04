"""Router de importación y actualización de Excel y CSV (CU-01, CU-10)."""
import os
import tempfile
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile, HTTPException, BackgroundTasks
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Licitacion, FicheroExcel, Config
from app.services.excel_parser import parse_excel
from app.services.csv_parser import parse_csv
from app.services.ai_service import calcular_idoneidad, generar_abreviado

router = APIRouter()


def _get_config(db: Session) -> Config:
    cfg = db.query(Config).first()
    if not cfg:
        cfg = Config()
        db.add(cfg)
        db.commit()
        db.refresh(cfg)
    return cfg


def _process_licitaciones_ai(db: Session, licitaciones: list, cfg: Config):
    """Calcula idoneidad y abreviado con IA para cada licitación."""
    if not getattr(cfg, 'calcular_idoneidad_import', True):
        for lic in licitaciones:
            lic.idoneidad_categoria = "no calculado"
            lic.abreviado = (lic.titulo or "")[:150]
        return

    api_key = cfg.gemini_api_key or os.getenv("GEMINI_API_KEY", "")
    model = cfg.gemini_model or "gemini-1.5-flash"
    desc = cfg.empresa_descripcion or ""

    for lic in licitaciones:
        if lic.titulo and api_key and desc:
            try:
                score, cat = calcular_idoneidad(desc, lic.titulo, api_key, model)
                lic.idoneidad_score = score
                lic.idoneidad_categoria = cat
            except Exception:
                lic.idoneidad_categoria = "baja"
            try:
                lic.abreviado = generar_abreviado(lic.titulo, api_key, model)
            except Exception:
                lic.abreviado = (lic.titulo or "")[:150]
        else:
            lic.abreviado = (lic.titulo or "")[:150]


def _parse_file(file: UploadFile) -> list:
    """Parsea Excel o CSV según extensión."""
    filename = (file.filename or "").lower()
    content = file.file.read()

    if filename.endswith((".csv",)):
        return parse_csv(content)
    if filename.endswith((".xlsx", ".xls")):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            rows = parse_excel(tmp_path)
        finally:
            os.unlink(tmp_path)
        return rows
    raise HTTPException(400, "Formato no soportado. Use Excel (.xlsx, .xls) o CSV (.csv con separador ';').")


@router.post("/import")
def import_excel(
    file: UploadFile,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """Importar nuevo Excel o CSV (CU-01). CSV: separador ';', columnas en fila 1."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "Debe subir un archivo Excel (.xlsx, .xls) o CSV (.csv)")

    try:
        rows = _parse_file(file)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Error al parsear archivo: {e}")

    if not rows:
        raise HTTPException(400, "El Excel no contiene filas válidas")

    cfg = _get_config(db)
    fichero = FicheroExcel(
        nombre_fichero=file.filename or "import.xlsx",
        fecha_carga=datetime.utcnow(),
        total_licitaciones=len(rows),
    )
    db.add(fichero)
    db.commit()
    db.refresh(fichero)

    licitaciones = []
    for r in rows:
        lic = Licitacion(
            fichero_id=fichero.id,
            tipo=r.get("tipo", ""),
            lugar=r.get("lugar", ""),
            organismo=r.get("organismo", ""),
            titulo=r.get("titulo", ""),
            abreviado=r.get("abreviado", ""),
            expediente=r.get("expediente", ""),
            fecha_licitacion=r.get("fecha_licitacion"),
            fecha_limite=r.get("fecha_limite"),
            hora_limite=r.get("hora_limite", ""),
            importe=r.get("importe", 0),
            url=r.get("url", ""),
            estado_decision="pendiente revisión idoneidad",
            estado_scraping="scraping_pendiente",
        )
        db.add(lic)
        licitaciones.append(lic)

    db.commit()
    # Calcular idoneidad en foreground para MVP (puede ser lento)
    _process_licitaciones_ai(db, licitaciones, cfg)
    db.commit()

    return {
        "fichero_id": fichero.id,
        "nombre": fichero.nombre_fichero,
        "total_importadas": len(rows),
    }


@router.post("/update")
def update_from_excel(
    file: UploadFile,
    db: Session = Depends(get_db),
):
    """Actualizar desde nuevo Excel o CSV (CU-10) - merge por expediente."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "Debe subir un archivo Excel (.xlsx, .xls) o CSV (.csv)")

    try:
        rows = _parse_file(file)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Error al parsear archivo: {e}")

    cfg = _get_config(db)
    fichero = FicheroExcel(
        nombre_fichero=file.filename or "update.xlsx",
        fecha_carga=datetime.utcnow(),
        total_licitaciones=0,
    )
    db.add(fichero)
    db.commit()
    db.refresh(fichero)

    nuevas = 0
    actualizadas = 0
    for r in rows:
        expediente = (r.get("expediente") or "").strip()
        existing = db.query(Licitacion).filter(Licitacion.expediente == expediente).first() if expediente else None

        if existing:
            existing.tipo = r.get("tipo", existing.tipo)
            existing.lugar = r.get("lugar", existing.lugar)
            existing.organismo = r.get("organismo", existing.organismo)
            existing.titulo = r.get("titulo", existing.titulo)
            existing.fecha_licitacion = r.get("fecha_licitacion") or existing.fecha_licitacion
            existing.fecha_limite = r.get("fecha_limite") or existing.fecha_limite
            existing.hora_limite = r.get("hora_limite", existing.hora_limite)
            existing.importe = r.get("importe", existing.importe)
            existing.url = r.get("url", existing.url)
            actualizadas += 1
        else:
            lic = Licitacion(
                fichero_id=fichero.id,
                tipo=r.get("tipo", ""),
                lugar=r.get("lugar", ""),
                organismo=r.get("organismo", ""),
                titulo=r.get("titulo", ""),
                abreviado="",
                expediente=expediente,
                fecha_licitacion=r.get("fecha_licitacion"),
                fecha_limite=r.get("fecha_limite"),
                hora_limite=r.get("hora_limite", ""),
                importe=r.get("importe", 0),
                url=r.get("url", ""),
                estado_decision="pendiente revisión idoneidad",
                estado_scraping="scraping_pendiente",
            )
            db.add(lic)
            nuevas += 1

    # Calcular idoneidad para nuevas
    nuevas_lics = [l for l in db.query(Licitacion).filter(Licitacion.fichero_id == fichero.id).all()]
    _process_licitaciones_ai(db, nuevas_lics, cfg)
    fichero.total_licitaciones = nuevas + actualizadas
    db.commit()

    return {
        "fichero_id": fichero.id,
        "nuevas": nuevas,
        "actualizadas": actualizadas,
    }


class ExportRequest(BaseModel):
    licitacion_ids: list[int]


@router.post("/export")
def export_licitaciones(
    body: ExportRequest,
    db: Session = Depends(get_db),
):
    """Exportar licitaciones seleccionadas a Excel (CU-05)."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl

    ids = body.licitacion_ids
    licitaciones = db.query(Licitacion).filter(Licitacion.id.in_(ids)).all()
    if not licitaciones:
        raise HTTPException(400, "No hay licitaciones seleccionadas")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Licitaciones"
    headers = [
        "ID", "Expediente", "Título", "Abreviado", "Organismo", "Lugar", "Importe",
        "Fecha límite", "URL", "Idoneidad", "Estado decisión", "Interesa", "Notas",
        "Riesgos", "Entregables",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)

    for row, lic in enumerate(licitaciones, 2):
        import json
        riesgos = json.loads(lic.riesgos_json or "[]") if lic.riesgos_json else []
        entregables = json.loads(lic.entregables_json or "[]") if lic.entregables_json else []
        ws.cell(row, 1, lic.id)
        ws.cell(row, 2, lic.expediente)
        ws.cell(row, 3, lic.titulo)
        ws.cell(row, 4, lic.abreviado)
        ws.cell(row, 5, lic.organismo)
        ws.cell(row, 6, lic.lugar)
        ws.cell(row, 7, lic.importe)
        ws.cell(row, 8, lic.fecha_limite.isoformat() if lic.fecha_limite else "")
        ws.cell(row, 9, lic.url)
        ws.cell(row, 10, lic.idoneidad_categoria)
        ws.cell(row, 11, lic.estado_decision)
        ws.cell(row, 12, "Sí" if lic.interesa else "No")
        ws.cell(row, 13, lic.notas)
        ws.cell(row, 14, "; ".join(riesgos) if isinstance(riesgos, list) else str(riesgos))
        ws.cell(row, 15, "; ".join(entregables) if isinstance(entregables, list) else str(entregables))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=licitaciones_export.xlsx"},
    )
